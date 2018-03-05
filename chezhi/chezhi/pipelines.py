# -*- coding: utf-8 -*-

import csv 
import os 
from openpyxl import Workbook
import time


class CsvPipeline(object):
        
    def process_item(self, item, spider):
        filename = os.getcwd() + '\\data\\%s_%s.csv' % (spider.name, item['updatetime'])
        if not os.path.exists(filename):
            with open(filename, 'a', newline='', encoding='utf-8') as csvfile:
                f = csv.writer(csvfile)
                f.writerow(sorted(item.keys()))
        with open(filename, 'a', newline='', encoding='utf-8') as csvfile:
            f = csv.writer(csvfile)
            f.writerow([item[key] for key in sorted(item.keys())])
            
class ExcelPipeline(object):
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.colTitle = []
        self.row = 2
        self.timestamp = str(int(time.time()))
        
    def process_item(self, item, spider):
        newColTitle = list(set(item.keys()).difference(set(self.colTitle))) #新列头
        self.colTitle.extend(newColTitle)
        
        for k,v in item.items():
            if k in newColTitle:
                self.ws.cell(
                    row = 1,
                    column = self.colTitle.index(k) + 1,
                    value = k
                )
            self.ws.cell(
                row = self.row,
                column = self.colTitle.index(k) + 1,
                value = v
            )
        self.row += 1
        
    def close_spider(self, spider):
        self.wb.save('.\\data\\{}_{}.xlsx'.format(spider.name, self.timestamp))   