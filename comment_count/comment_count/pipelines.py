# -*- coding: utf-8 -*-

import csv 
import os 
from openpyxl import Workbook
import time
import pandas as pd

class CommentCountPipeline(object):
        
    def process_item(self, item, spider):
        filename = os.getcwd() + '\\data\\%s_%s.csv' % (spider.name, item['updatetime'])
        if not os.path.exists(filename):
            with open(filename, 'a', newline='', encoding='utf-8') as csvfile:
                f = csv.writer(csvfile)
                f.writerow(sorted(item.keys()))
        with open(filename, 'a', newline='', encoding='utf-8') as csvfile:
            f = csv.writer(csvfile)
            f.writerow([item[key] for key in sorted(item.keys())])
            
class SkuPipeline(object):
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.colTitle = []
        self.row = 2
        self.timestamp = str(int(time.time()))
        
    def process_item(self, item, spider):
        tmpItem = {k:v for k,v in item.items() if k!='基本信息'}
        for tmp_sku in item.get('基本信息',[]):
            try: tmpItem.update(tmp_sku)
            except: pass
        newColTitle = list(set(tmpItem.keys()).difference(set(self.colTitle))) #新列头
        self.colTitle.extend(newColTitle)
        
        for k,v in tmpItem.items():
            if k in newColTitle:
                self.ws.cell(row=1,column=self.colTitle.index(k)+1,value = k)
            self.ws.cell(row=self.row,column=self.colTitle.index(k)+1,value=v)
        self.row += 1
        
    def close_spider(self, spider):
        filename = '.\\data\\{}_{}.xlsx'.format(spider.name, self.timestamp)
        self.wb.save(filename)   
        self._deal_excel(filename)
    
    def _deal_excel(self,filename):
        """过滤掉excel表中非空值比例低于10%的列"""
        if os.path.exists(filename):
            df = pd.read_excel(filename)
            col_idx = list(df.columns)
            for idx in col_idx:
                col_data = list(df[idx])
                percent = len([i for i in col_data if str(i) != 'nan']) / len(col_data)
                if percent < 0.1:
                    del(df[idx])
            writer = pd.ExcelWriter(filename,
                        engine='xlsxwriter',
                        options={'strings_to_urls':False}
                    )
            df.to_excel(writer, index=False)
            writer.close()

    
class CommentCountPipeline(object):
        
    def process_item(self, item, spider):
        filename = os.getcwd() + '\\data\\%s_%s.csv' % (spider.name, item['updatetime'])
        if not os.path.exists(filename):
            with open(filename, 'a', newline='', encoding='utf-8') as csvfile:
                f = csv.writer(csvfile)
                f.writerow(sorted(item.keys()))
        with open(filename, 'a', newline='', encoding='utf-8') as csvfile:
            f = csv.writer(csvfile)
            f.writerow([item[key] for key in sorted(item.keys())])